import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from apps.catalogos.models import CatalogoCuenta, TipoCuenta
from apps.estados.models import TipoEstadoFinanciero


def generar_plantilla_balance_general(catalogo):
    """
    Genera un archivo Excel de plantilla para cargar Balance General.
    Incluye solo cuentas de tipo ACTIVO, PASIVO y PATRIMONIO.
    
    Args:
        catalogo: Instancia de CatalogoCuenta
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance General"
    
    # Encabezados: todas las columnas del catálogo + monto
    headers = ['Código de la cuenta', 'Nombre de la cuenta', 'Tipo de cuenta', 'Monto']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filtrar cuentas: solo ACTIVO, PASIVO, PATRIMONIO
    tipos_permitidos = [TipoCuenta.ACTIVO, TipoCuenta.PASIVO, TipoCuenta.PATRIMONIO]
    cuentas = catalogo.cuentas.filter(tipo__in=tipos_permitidos).order_by('codigo', 'nombre')
    
    # Agregar datos de las cuentas
    for row_num, cuenta in enumerate(cuentas, 2):
        ws.cell(row=row_num, column=1, value=cuenta.codigo)
        ws.cell(row=row_num, column=2, value=cuenta.nombre)
        ws.cell(row=row_num, column=3, value=cuenta.get_tipo_display())
        ws.cell(row=row_num, column=4, value=0.00)  # Valor por defecto para monto
        
        # Formatear celda de monto como número con 2 decimales
        monto_cell = ws.cell(row=row_num, column=4)
        monto_cell.number_format = '#,##0.00'
        monto_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Ajustar ancho de columnas
    column_widths = [25, 40, 25, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Crear hoja de ayuda
    ws_ayuda = wb.create_sheet("Ayuda")
    ws_ayuda['A1'] = 'Instrucciones:'
    ws_ayuda['A1'].font = Font(bold=True)
    ws_ayuda['A2'] = '1. Complete la columna "Monto" con los valores correspondientes.'
    ws_ayuda['A3'] = '2. No modifique las columnas Código, Nombre o Tipo de cuenta.'
    ws_ayuda['A4'] = '3. Los montos deben ser números (pueden incluir decimales).'
    ws_ayuda['A5'] = '4. Si una cuenta no aplica, deje el monto en 0.00.'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generar_plantilla_estado_resultados(catalogo):
    """
    Genera un archivo Excel de plantilla para cargar Estado de Resultados.
    Incluye solo cuentas de tipo INGRESO, GASTO y RESULTADO.
    
    Args:
        catalogo: Instancia de CatalogoCuenta
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"
    
    # Encabezados: todas las columnas del catálogo + monto
    headers = ['Código de la cuenta', 'Nombre de la cuenta', 'Tipo de cuenta', 'Monto']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filtrar cuentas: solo INGRESO, GASTO, RESULTADO
    tipos_permitidos = [TipoCuenta.INGRESO, TipoCuenta.GASTO, TipoCuenta.RESULTADO]
    cuentas = catalogo.cuentas.filter(tipo__in=tipos_permitidos).order_by('codigo', 'nombre')
    
    # Agregar datos de las cuentas
    for row_num, cuenta in enumerate(cuentas, 2):
        ws.cell(row=row_num, column=1, value=cuenta.codigo)
        ws.cell(row=row_num, column=2, value=cuenta.nombre)
        ws.cell(row=row_num, column=3, value=cuenta.get_tipo_display())
        ws.cell(row=row_num, column=4, value=0.00)  # Valor por defecto para monto
        
        # Formatear celda de monto como número con 2 decimales
        monto_cell = ws.cell(row=row_num, column=4)
        monto_cell.number_format = '#,##0.00'
        monto_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Ajustar ancho de columnas
    column_widths = [25, 40, 25, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Crear hoja de ayuda
    ws_ayuda = wb.create_sheet("Ayuda")
    ws_ayuda['A1'] = 'Instrucciones:'
    ws_ayuda['A1'].font = Font(bold=True)
    ws_ayuda['A2'] = '1. Complete la columna "Monto" con los valores correspondientes.'
    ws_ayuda['A3'] = '2. No modifique las columnas Código, Nombre o Tipo de cuenta.'
    ws_ayuda['A4'] = '3. Los montos deben ser números (pueden incluir decimales).'
    ws_ayuda['A5'] = '4. Si una cuenta no aplica, deje el monto en 0.00.'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generar_plantilla_excel_estado(catalogo, tipo_estado):
    """
    Genera un archivo Excel de plantilla según el tipo de estado financiero.
    
    Args:
        catalogo: Instancia de CatalogoCuenta
        tipo_estado: TipoEstadoFinanciero (BALANCE_GENERAL o ESTADO_RESULTADOS)
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    if tipo_estado == TipoEstadoFinanciero.BALANCE_GENERAL:
        return generar_plantilla_balance_general(catalogo)
    elif tipo_estado == TipoEstadoFinanciero.ESTADO_RESULTADOS:
        return generar_plantilla_estado_resultados(catalogo)
    else:
        raise ValueError(f'Tipo de estado financiero no válido: {tipo_estado}')

