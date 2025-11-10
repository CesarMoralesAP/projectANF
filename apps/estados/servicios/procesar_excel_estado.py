import openpyxl
from decimal import Decimal, InvalidOperation
from apps.estados.models import EstadoFinanciero, ItemEstadoFinanciero, TipoEstadoFinanciero
from apps.empresas.models import Empresa
from apps.catalogos.models import CatalogoCuenta, CuentaContable, TipoCuenta
from django.db import transaction


def procesar_excel_estado(archivo, empresa, año, tipo_estado):
    """
    Procesa un archivo Excel y crea/actualiza un estado financiero.
    
    Args:
        archivo: Archivo Excel cargado
        empresa: Instancia de Empresa
        año: Año del estado financiero
        tipo_estado: TipoEstadoFinanciero (BALANCE_GENERAL o ESTADO_RESULTADOS)
    
    Returns:
        tuple: (éxito: bool, errores: list, items_procesados: int)
    """
    errores = []
    items_procesados = 0
    éxito = False
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Validar encabezados (normalizar para comparación)
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        headers_esperados = ['Código de la cuenta', 'Nombre de la cuenta', 'Tipo de cuenta', 'Monto']
        
        # Comparar normalizando (case-insensitive y sin espacios extras)
        headers_normalizados = [h.strip().lower() for h in headers]
        headers_esperados_normalizados = [h.strip().lower() for h in headers_esperados]
        
        if headers_normalizados != headers_esperados_normalizados:
            errores.append(
                f'Los encabezados no coinciden. Se esperaban: {", ".join(headers_esperados)}. '
                f'Se encontraron: {", ".join(headers)}'
            )
            return False, errores, 0
        
        # Obtener o crear catálogo de cuentas
        try:
            catalogo = CatalogoCuenta.objects.get(empresa=empresa)
        except CatalogoCuenta.DoesNotExist:
            errores.append(f'La empresa "{empresa.nombre}" no tiene un catálogo de cuentas configurado.')
            return False, errores, 0
        
        # Validar tipos de cuenta según el tipo de estado
        if tipo_estado == TipoEstadoFinanciero.BALANCE_GENERAL:
            tipos_permitidos = [TipoCuenta.ACTIVO, TipoCuenta.PASIVO, TipoCuenta.PATRIMONIO]
        elif tipo_estado == TipoEstadoFinanciero.ESTADO_RESULTADOS:
            tipos_permitidos = [TipoCuenta.INGRESO, TipoCuenta.GASTO, TipoCuenta.RESULTADO]
        else:
            errores.append(f'Tipo de estado financiero no válido: {tipo_estado}')
            return False, errores, 0
        
        # Procesar con transacción atómica
        with transaction.atomic():
            # Obtener o crear estado financiero
            estado_financiero, creado = EstadoFinanciero.objects.get_or_create(
                empresa=empresa,
                año=año,
                tipo=tipo_estado,
                defaults={}
            )
            
            # Si ya existía, eliminar items existentes para reemplazarlos
            if not creado:
                ItemEstadoFinanciero.objects.filter(estado_financiero=estado_financiero).delete()
            
            # Procesar filas (empezar desde la fila 2, ya que la 1 son los encabezados)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Saltar filas vacías
                if not any(row):
                    continue
                
                codigo = str(row[0]).strip() if row[0] else None
                nombre = str(row[1]).strip() if row[1] else None
                tipo_cuenta_str = str(row[2]).strip() if row[2] else None
                monto_str = row[3] if row[3] is not None else None
                
                # Validar campos requeridos
                if not codigo:
                    errores.append(f'Fila {row_num}: Falta el código de la cuenta')
                    continue
                
                if not nombre:
                    errores.append(f'Fila {row_num}: Falta el nombre de la cuenta')
                    continue
                
                # Validar que la cuenta existe en el catálogo
                try:
                    cuenta = CuentaContable.objects.get(
                        catalogo=catalogo,
                        codigo=codigo
                    )
                except CuentaContable.DoesNotExist:
                    errores.append(
                        f'Fila {row_num}: La cuenta con código "{codigo}" no existe en el catálogo de la empresa.'
                    )
                    continue
                
                # Validar que el tipo de cuenta coincide (comparar normalizado)
                # El tipo_cuenta_str puede venir como "Activo" (display) o "ACTIVO" (valor)
                tipo_cuenta_normalizado = tipo_cuenta_str.upper().strip()
                cuenta_tipo_normalizado = cuenta.tipo.upper().strip()
                
                # Verificar si el tipo coincide (puede venir como valor o como display)
                tipos_coinciden = (
                    cuenta_tipo_normalizado == tipo_cuenta_normalizado or
                    cuenta.get_tipo_display().upper().strip() == tipo_cuenta_normalizado
                )
                
                if not tipos_coinciden:
                    errores.append(
                        f'Fila {row_num}: El tipo de cuenta no coincide. '
                        f'En Excel: "{tipo_cuenta_str}", En catálogo: "{cuenta.get_tipo_display()}" (valor: {cuenta.tipo})'
                    )
                    continue
                
                # Validar que el tipo de cuenta es permitido para este estado
                if cuenta.tipo not in tipos_permitidos:
                    errores.append(
                        f'Fila {row_num}: La cuenta "{codigo}" de tipo "{cuenta.get_tipo_display()}" '
                        f'no es válida para un {estado_financiero.get_tipo_display()}.'
                    )
                    continue
                
                # Validar y convertir monto
                if monto_str is None or monto_str == '':
                    monto = Decimal('0.00')
                else:
                    try:
                        # Convertir a Decimal
                        if isinstance(monto_str, (int, float)):
                            monto = Decimal(str(monto_str))
                        elif isinstance(monto_str, str):
                            # Limpiar el string (remover espacios, comas como separadores de miles)
                            # Pero mantener el punto como separador decimal
                            monto_str_limpio = monto_str.strip()
                            # Si tiene coma, puede ser separador de miles o decimal
                            # Si tiene punto, el punto es decimal
                            if '.' in monto_str_limpio and ',' in monto_str_limpio:
                                # Formato tipo "1.234,56" o "1,234.56"
                                if monto_str_limpio.rindex('.') > monto_str_limpio.rindex(','):
                                    # Punto después de coma: "1,234.56" (formato inglés)
                                    monto_str_limpio = monto_str_limpio.replace(',', '')
                                else:
                                    # Coma después de punto: "1.234,56" (formato europeo)
                                    monto_str_limpio = monto_str_limpio.replace('.', '').replace(',', '.')
                            elif ',' in monto_str_limpio:
                                # Solo coma: puede ser decimal o miles
                                # Si hay más de 3 dígitos después de la coma, probablemente es decimal
                                partes = monto_str_limpio.split(',')
                                if len(partes) == 2 and len(partes[1]) <= 2:
                                    # Coma como separador decimal
                                    monto_str_limpio = monto_str_limpio.replace(',', '.')
                                else:
                                    # Coma como separador de miles
                                    monto_str_limpio = monto_str_limpio.replace(',', '')
                            
                            monto = Decimal(monto_str_limpio)
                        else:
                            monto = Decimal('0.00')
                    except (InvalidOperation, ValueError, Exception) as e:
                        errores.append(
                            f'Fila {row_num}: El monto "{monto_str}" no es un número válido. Error: {str(e)}'
                        )
                        continue
                
                # Crear item del estado financiero (solo si el monto es diferente de 0 o si queremos guardar todos)
                # Por ahora, guardamos todos los items, incluso si el monto es 0
                try:
                    ItemEstadoFinanciero.objects.create(
                        estado_financiero=estado_financiero,
                        cuenta_contable=cuenta,
                        monto=monto
                    )
                    items_procesados += 1
                except Exception as e:
                    errores.append(f'Fila {row_num}: Error al crear item - {str(e)}')
                    import traceback
                    print(f"Error al crear item en fila {row_num}: {traceback.format_exc()}")
            
            éxito = len(errores) == 0
    
    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        errores.append(f'Error al procesar el archivo: {str(e)}')
        # Agregar detalles del error para debugging
        print(f"Error al procesar Excel: {error_detalle}")
        éxito = False
    
    # Considerar éxito si se procesaron items, incluso si hubo algunos errores
    # El éxito total solo será True si no hubo errores
    # Pero retornamos los items procesados para que la vista pueda mostrar un mensaje apropiado
    return éxito, errores, items_procesados

