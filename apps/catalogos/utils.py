import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from .models import TipoCuenta, CuentaContable, CatalogoCuenta
from django.core.exceptions import ValidationError


def generar_plantilla_excel():
    """
    Genera un archivo Excel de plantilla para cargar cuentas contables.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuentas Contables"
    
    # Encabezados
    headers = ['Código de la cuenta', 'Nombre de la cuenta', 'Tipo de cuenta']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ejemplo de datos
    ejemplos = [
        ['0101', 'Activos corrientes', 'ACTIVO'],
        ['0102', 'Efectivo y equivalentes', 'ACTIVO'],
        ['0201', 'Pasivos corrientes', 'PASIVO'],
        ['0301', 'Capital social', 'PATRIMONIO'],
        ['0401', 'Ingresos operativos', 'INGRESO'],
        ['0501', 'Gastos operativos', 'GASTO'],
        ['0601', 'Ganancias', 'RESULTADO'],

    ]
    
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, value in enumerate(ejemplo, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar ancho de columnas
    column_widths = [25, 40, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Crear validación de datos para la columna de tipo de cuenta
    tipo_choices = [tipo[0] for tipo in TipoCuenta.choices]
    tipo_validation = f'"{",".join(tipo_choices)}"'
    
    # Crear hoja de ayuda con los tipos válidos
    ws_ayuda = wb.create_sheet("Ayuda")
    ws_ayuda['A1'] = 'Tipos de cuenta válidos:'
    ws_ayuda['A1'].font = Font(bold=True)
    for idx, (valor, nombre) in enumerate(TipoCuenta.choices, 2):
        ws_ayuda[f'A{idx}'] = f'{valor} - {nombre}'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def procesar_excel(archivo, catalogo):
    """
    Procesa un archivo Excel y reemplaza todas las cuentas contables del catálogo.
    Primero elimina todas las cuentas existentes y luego crea las nuevas.
    
    Args:
        archivo: Archivo Excel cargado
        catalogo: Instancia de CatalogoCuenta
    
    Returns:
        tuple: (cuentas_creadas, errores)
    """
    errores = []
    cuentas_creadas = 0
    
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Validar encabezados
        headers = [cell.value for cell in ws[1]]
        headers_esperados = ['Código de la cuenta', 'Nombre de la cuenta', 'Tipo de cuenta']
        
        if headers != headers_esperados:
            errores.append(
                f'Los encabezados no coinciden. Se esperaban: {", ".join(headers_esperados)}'
            )
            return cuentas_creadas, errores
        
        # Eliminar todas las cuentas existentes del catálogo
        CuentaContable.objects.filter(catalogo=catalogo).delete()
        
        # Procesar filas (empezar desde la fila 2, ya que la 1 son los encabezados)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas vacías
            if not any(row):
                continue
            
            codigo = str(row[0]).strip() if row[0] else None
            nombre = str(row[1]).strip() if row[1] else None
            tipo = str(row[2]).strip() if row[2] else None
            
            # Validar campos requeridos
            if not codigo or not nombre or not tipo:
                errores.append(f'Fila {row_num}: Faltan campos requeridos')
                continue
            
            # Validar tipo de cuenta
            tipos_validos = [tipo[0] for tipo in TipoCuenta.choices]
            if tipo not in tipos_validos:
                errores.append(
                    f'Fila {row_num}: Tipo de cuenta "{tipo}" no válido. '
                    f'Tipos válidos: {", ".join(tipos_validos)}'
                )
                continue
            
            # Crear cuenta
            try:
                CuentaContable.objects.create(
                    catalogo=catalogo,
                    codigo=codigo,
                    nombre=nombre,
                    tipo=tipo
                )
                cuentas_creadas += 1
            except Exception as e:
                errores.append(f'Fila {row_num}: Error al crear cuenta - {str(e)}')
    
    except Exception as e:
        errores.append(f'Error al procesar el archivo: {str(e)}')
    
    return cuentas_creadas, errores

